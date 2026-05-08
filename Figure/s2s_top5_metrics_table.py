import argparse
import re
from pathlib import Path

import numpy as np
from openpyxl import load_workbook


def safe_float(value):
    if value is None:
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def format_receipts(value):
    value = float(value)
    if value >= 1e9:
        return f"{value / 1e9:.2f}B"
    if value >= 1e6:
        return f"{value / 1e6:.2f}M"
    if value >= 1e3:
        return f"{value / 1e3:.2f}K"
    return f"{value:.0f}"


def compute_metrics(matrix):
    matrix = np.asarray(matrix, dtype=float)
    if matrix.shape[0] != matrix.shape[1]:
        raise ValueError("Matrix M must be square.")

    n = matrix.shape[0]
    total = matrix.sum()
    diagonal = np.trace(matrix)
    cross = total - diagonal

    csr = (cross / total) if total > 0 else 0.0

    out_load = matrix.sum(axis=1)
    in_load = matrix.sum(axis=0)
    load = out_load + in_load
    cv_load = (np.std(load) / np.mean(load)) if np.mean(load) > 0 else 0.0

    asym_num = 0.0
    asym_den = 0.0
    for i in range(n):
        for j in range(i + 1, n):
            a = matrix[i, j]
            b = matrix[j, i]
            asym_num += abs(a - b)
            asym_den += (a + b)
    asym = (asym_num / asym_den) if asym_den > 0 else 0.0

    hhi_values = []
    for i in range(n):
        if out_load[i] <= 0:
            continue
        p = matrix[i, :] / out_load[i]
        hhi_values.append(float(np.sum(p ** 2)))
    mean_hhi_out = float(np.mean(hhi_values)) if hhi_values else 0.0

    off_diag = matrix.copy()
    np.fill_diagonal(off_diag, 0.0)
    top_pair_share = (off_diag.max() / cross) if cross > 0 else 0.0
    if cross > 0:
        top_pair_row, top_pair_col = np.unravel_index(np.argmax(off_diag), off_diag.shape)
        top_pair = f"S{top_pair_row}->S{top_pair_col}"
    else:
        top_pair = "-"

    return {
        "receipts_total": float(total),
        "csr": float(csr),
        "cv_load": float(cv_load),
        "asym": float(asym),
        "hhi_out_mean": float(mean_hhi_out),
        "top_pair": top_pair,
        "top_pair_share": float(top_pair_share),
    }


def _is_shard_label(value):
    return isinstance(value, str) and re.match(r"^\s*S\d+\s*$", value, re.IGNORECASE) is not None


def find_matrix_block(worksheet, shard_count):
    marker = re.compile(rf"^\s*{shard_count}\s*shards?\s*$", re.IGNORECASE)

    for row in worksheet.iter_rows(min_row=1, max_row=150, min_col=1, max_col=250):
        for cell in row:
            if isinstance(cell.value, str) and marker.match(cell.value.strip()):
                anchor_row, anchor_col = cell.row, cell.column

                # Validate that the row/column labels near this anchor look like shard labels.
                col_labels = [
                    worksheet.cell(row=anchor_row, column=anchor_col + 1 + j).value
                    for j in range(shard_count)
                ]
                row_labels = [
                    worksheet.cell(row=anchor_row + 1 + i, column=anchor_col).value
                    for i in range(shard_count)
                ]
                if not all(_is_shard_label(v) for v in col_labels):
                    continue
                if not all(_is_shard_label(v) for v in row_labels):
                    continue

                matrix = np.zeros((shard_count, shard_count), dtype=float)
                for i in range(shard_count):
                    for j in range(shard_count):
                        matrix[i, j] = safe_float(
                            worksheet.cell(row=anchor_row + 1 + i, column=anchor_col + 1 + j).value
                        )
                return matrix

    return None


def extract_matrices(workbook_path, shard_counts, sheet_name="S2S (Receipt)"):
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    matrices = {}

    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")

    worksheet = workbook[sheet_name]

    for shard_count in shard_counts:
        matrix = find_matrix_block(worksheet, shard_count)
        if matrix is not None:
            matrices[shard_count] = matrix

    return matrices


def build_latex(rows):
    lines = [
        r"\begin{table}[t]",
        r"\centering",
        r"\small",
        r"\renewcommand{\arraystretch}{1.1}",
        r"\caption{Flow-based shard interaction metrics derived from the matrix $M$.}",
        r"\label{tab:shard_interaction_final}",
        r"\begin{tabular}{c r r r r r r}",
        r"\toprule",
        r"\# Shards & \# Receipts & $\overline{C}$ & CV$(L)$ & Asym & mean HHI$_{out}$ & Top Pair (Share) \\",
        r"\midrule",
    ]

    for row in sorted(rows, key=lambda x: x["shard"]):
        lines.append(
            f"{row['shard']} & "
            f"{format_receipts(row['receipts_total'])} & "
            f"{row['csr']:.2f} & "
            f"{row['cv_load']:.2f} & "
            f"{row['asym']:.2f} & "
            f"{row['hhi_out_mean']:.2f} & "
            f"{row['top_pair']} ({row['top_pair_share']:.2f}) \\\\" 
        )

    lines.extend(
        [
            r"\bottomrule",
            r"\end{tabular}",
            r"\end{table}",
        ]
    )
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(
        description="Compute top 5 S2S-only metrics and print LaTeX table."
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=Path(__file__).with_name("Fig_all_new.xlsx"),
        help="Path to workbook containing S2S matrices.",
    )
    parser.add_argument(
        "--shards",
        type=int,
        nargs="+",
        default=[4, 5, 6, 7, 8, 9],
        help="Shard counts to extract from workbook.",
    )
    parser.add_argument(
        "--save-tex",
        type=Path,
        default=None,
        help="Optional output .tex path.",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default="S2S (Receipt)",
        help="Worksheet name containing shard-to-shard matrices.",
    )
    args = parser.parse_args()

    matrices = extract_matrices(args.xlsx, args.shards, sheet_name=args.sheet)

    rows = []
    missing = []
    for shard in args.shards:
        if shard not in matrices:
            missing.append(shard)
            continue

        metrics = compute_metrics(matrices[shard])
        rows.append({"shard": shard, **metrics})

    latex = build_latex(rows)
    print(latex)

    if args.save_tex is not None:
        args.save_tex.write_text(latex, encoding="utf-8")
        print(f"\nSaved LaTeX table to: {args.save_tex}")

    if missing:
        print(f"\nMissing shard blocks: {missing}")


if __name__ == "__main__":
    main()
